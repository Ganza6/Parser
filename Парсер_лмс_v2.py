import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from string import ascii_uppercase

all_correct = 0 
all_incorrect = 0
all_relatively = 0
all_page_number = 0
page_number = 0
start_time = time.time()
correct = {}
incorrect = {}
relatively = {}
broken_acc = {}
name = ''


def get_html(session, url):
        r = session.get(url)
        return r.text


def save(size):
    wb = Workbook()
    str1 = wb.active
    str1.title = 'Правильные'
    str2 = wb.create_sheet('Неправильные')
    str3 = wb.create_sheet('Частично правильные и составные')
    str1['A1'] = 'Вопрос'
    str1['B1'] = 'Ответ'
    str2['A1'] = 'Вопрос'
    str2['B1'] = 'Ответ'
    str3['A1'] = 'Вопрос'
    str3['B1'] = 'Ответ'
    str3['C1'] = 'Баллы'
    for column in ascii_uppercase:
        if column == 'A':
                str1.column_dimensions[column].width = 50
                str3.column_dimensions[column].width = 50
                str2.column_dimensions[column].width = 50
        if column == 'B':
                str1.column_dimensions[column].width = 50
                str3.column_dimensions[column].width = 50
                str2.column_dimensions[column].width = 50
    for i in correct:
        string = ', '.join(correct[i])
        str1.append((i,string))
    for i in incorrect:
        string = ', '.join(incorrect[i])  # массив неправильных ответов
        str2.append((i,string))
    key_sorted = sorted(relatively)
    for i in key_sorted:
        str3.append((i[0],i[1],relatively[i]))
    time_now = a = time.ctime(time.time()).replace(":", "`")
    file_name = str(size) + ' ' + name+'. '+time_now+".xlsx"
    wb.save(file_name)


def get_acc_base():
    file = open('lms_accounts.txt', 'r')
    mass = []
    for acc in file:
        login = acc[:6]
        password = acc[7:11]
        mass.append((login,password))
    return mass


def show_info():
    global all_correct, all_incorrect, all_relatively, all_page_number
    print('\nНазвание теста "%s"\n' % name[9:])
    print("Количество проверенных попыток %d" % page_number)
    all_correct += len(correct)
    all_incorrect += len(incorrect)
    all_relatively += len(relatively)
    all_page_number += page_number
    print("Количество правильных ответов %d\nКоличество неправильных ответов %d\n"
          "Количество частично правильных ответов %d\n\n" % (len(correct),len(incorrect),len(relatively)))


def all_info(all_size, broken_size, urls_number):
    print("\n\nОбщая статистика за сессию\n")
    print("Время работы программы %d секунд" % (time.time() - start_time))
    print("Проверено аккаунтов %d" % (all_size-broken_size))
    print("Количество аккаунтов, в которые не удалось войти %d" % broken_size)
    print("Количество тестов %d" % urls_number)
    print("Общее количество проверенных попыток %d" % all_page_number)
    print("Общее количество:\n правильных ответов %d\n неправильных ответов %d\n частично правильных ответов %d\n\n" % 
          (all_correct, all_incorrect, all_relatively))


def auth(session, login, password):
    url = 'http://lms.mai.ru/login/index.php'
    parameters = {'username': login,
                  'password': password,
                  'rememberusername': '1'
                  }
    session.post(url,parameters)


def get_url_tests(html):
    global page_number
    global name
    links = []
    soup = BeautifulSoup(html, 'lxml')
    try:
        name = soup.find(class_='main').get_text()
    except AttributeError:
        return 0
    title_with_ref = soup.find_all(title="Просмотр своих ответов этой попытки")
    for link in title_with_ref:
        page_number += 1
        links.append(link.get('href'))
    return links


def get_url_question(html):
    links = []
    soup = BeautifulSoup(html, 'lxml')
    title_with_ref = soup.find_all(class_="qnbutton complete free")
    for link in title_with_ref:
        links.append(link.get('href'))
    return links


def get_answer(html):
    soup = BeautifulSoup(html,'lxml')
    questions = soup.find_all(class_="que multichoice deferredfeedback complete")
    for question in questions:
        checked_answers = []
        all_answer = question.find_all(class_="r0") + question.find_all(class_="r1")
        text_question = question.find(class_='qtext').get_text()
        for answer in all_answer:
            if answer.find(checked='checked'):
                checked_answer = answer.find('label').get_text()
                if len(checked_answer) == 3:  # проверка ответ это фото или нет
                    img = answer.find('img')
                    checked_answers.append(img['src'])
                else:  # eсли это не фото
                    checked_answers.append(checked_answer[3:])
        result = question.find(class_ = 'grade').get_text()
        result = result[8:12]
        if result == '1,00':
            correct[text_question] = checked_answers
            incorrect.pop(text_question,1)
            relatively.pop(text_question,1)
        elif correct.get(text_question,0) == 0:
            if result == '0,00' and len(checked_answers) == 1:
                if incorrect.get(text_question,0):
                        incorrect[text_question] |= set(checked_answers)
                else:
                    incorrect[text_question] = set(checked_answers)
            else:
                relatively[(text_question,', '.join(checked_answers))] = result
    return 1


def check(html):
    soup = BeautifulSoup(html,'lxml')
    string = soup.find(class_="logininfo").get_text()
    if string.find('Вы зашли под именем') == 0:
        print(string[:-8])
        return 1
    else:
        print(string)
        return 0


def main():
    global correct, incorrect, relatively
    urls = ['http://lms.mai.ru/mod/quiz/view.php?id=19492','http://lms.mai.ru/mod/quiz/view.php?id=19491']
    urls_number = len(urls)
    acc_base = get_acc_base()
    for url in urls:
        correct, incorrect, relatively = {},{},{} #  обнуляем значения словарей для следующего теста
        page_number = 0
        for i in acc_base:
            session = requests.Session()
            login = i[0]
            password = i[1]
            auth(session,login,password)
            if check(get_html(session,url)) == 0:
                broken_acc[login]=password
                continue
            links_on_test = get_url_tests(get_html(session,url))
            if links_on_test:
                for link in links_on_test:
                    get_answer(get_html(session,link+'&showall=1'))
        show_info()
        size = len(correct)+len(incorrect)+len(relatively)
        if size>0:
            save(size)
    all_info(len(acc_base),len(broken_acc),urls_number)

if __name__ =='__main__':
    main()
